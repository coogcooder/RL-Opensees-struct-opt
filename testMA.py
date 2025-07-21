import copy
import time
import os
import numpy as np
import matplotlib.pyplot as plt
import torch
import torch.nn as nn
import torch.nn.functional as F
import torch.optim as optim
from torch.autograd import Variable
import openpyxl as excel
import traceback
import sys
sys.path.append("~/module")
###Original imports
from module import envmnt_lrn
from module import anly
from module import visualize
from module import NN
from module import parameters as para
####GPU
torch.set_default_tensor_type('torch.cuda.FloatTensor')
device = 'cuda' if torch.cuda.is_available() else 'cpu'
print(device)
###Environment expansion


    # Model
Qel = NN.NN1(para.obs_num1,para.acts_num1).to(device)   
Qdim= NN.NN2(para.obs_num2,para.acts_num2).to(device) 

env = envmnt_lrn.Env()
di=input()
num=input()
Qel.load_state_dict(torch.load('./output_images/learning_data'+str(di)+'/agentEL'+str(num)+'.pth'))   
Qdim.load_state_dict(torch.load('./output_images/learning_data'+str(di)+'/agentDIM'+str(num)+'.pth'))  
from openpyxl.styles import PatternFill
fill = PatternFill( patternType='solid', fgColor='B8CCE4', bgColor= 'B8CCE4')
wb=excel.Workbook()
ws=wb['Sheet']
sw=0
pp=[]
while sw<1:
   xSpan,ySpan,zSpan,lx,ly,load=int(input('xSpan')),int(input('ySpan')),int(input('zSpan')),int(input('lx'))*1000,int(input('ly'))*1000,int(input('load'))
   pp.append([xSpan,ySpan,zSpan,lx,ly,load])
   sw=int(input())
############TEST########

for condition in pp:
                print(condition)
            
                t=0                       
                done_t = False                           
                total_reward_t = 0     #      
                env.reset(condition) 
                print(env.memberz[env.margin:-env.margin,env.margin:-env.margin,env.margin:-env.margin])

                print('Stress ratio')
                for ii in range(env.zSpan):
                    print(np.round(env.stress_ratio[ii,:,:],2))
                print('COFX')
                for ii in range(env.zSpan):
                    print(np.round(env.cofX[ii,:,:],2))
                print('COFY')
                for ii in range(env.zSpan):
                    print(np.round(env.cofY[ii,:,:],2))

                print('Inter-story drift angle')  
                print(np.round(env.deflect,2))     

                print((str(env.step_count)+' steps completed').rjust(5, ' ')
                    ,' violation',str(np.round(env.violation,3)).rjust(6, ' ')
                    ,' volume',str(env.V).rjust(6, ' ')
                    )                              
                start=time.time()
                while not done_t :
                    Qs=env.value_scan_assets(Qel,device)
                    if env.flag_counter>0:
                        for i in range(env.flag_counter):
                            index = np.argmax(Qs)
                            Qs[index] =-100
                    index = np.argmax(Qs) 
                    if  t==1000 or np.max(Qs)==-100:break
                
                    index = np.argmax(Qs)
                    tnum_el=env.sort[index]
                    update_pos = index

                    list_dim=env.dim_assets(env.sort[index],env.memberz[tnum_el[0],tnum_el[1],tnum_el[2]])

                    if len(list_dim)==0:env.flag_counter +=1
                    else:
                        # 最適な行動を予測
                        pobs = anly.state_updateDIM(copy.copy(env.stress_ratio),copy.copy(env.cofX),copy.copy(env.cofY)
                                                    ,copy.copy(env.deflect),tnum_el,env.margin,env.arraymat,env.memberz)                    
                        pobs_ = np.array(pobs, dtype="float32").reshape((1, para.obs_num2))
                        pobs_ = torch.from_numpy(pobs_).to(device)
                        pacts = Qdim(pobs_).to('cpu').detach().numpy().copy()[0]
                        pacts=pacts[list_dim]
                        id=np.argmax(pacts)
                        pact=list_dim[id]                         
                        # 行動 
                        reward, _,member,dviolation,pact= env.step(update_pos,pact) 
                        total_reward_t += max([0,reward])    
                    
                        
                        sign='●'
                        if env.F==False or env.flag==True:
                            sign='　'    
                            env.memberz,env.stress_ratio,env.cofX,env.cofY,env.deflect,env.violation,env.V=(
                                copy.copy(env.result[0]),copy.copy(env.result[1]),copy.copy(env.result[2]),copy.copy(env.result[3])
                                        ,copy.copy(env.result[4]),copy.copy(env.result[5]),copy.copy(env.result[6]))               
                        else:env.flag_counter=0
                        t=t+1
        
                        ws['B'+str(env.step_count+1)]=env.step_count
                        ws['C'+str(env.step_count+1)]=total_reward_t
                        ws['D'+str(env.step_count+1)]=env.V
                        ws['E'+str(env.step_count+1)]=env.violation   
                        ws['F'+str(env.step_count+1)]=env.resultb[5] 
                        var=anly.variable_ex(env.memberz,envmnt_lrn.col_list,envmnt_lrn.beam_list,env.arraymat)
                        for num in range(len(var)):
                            ab=str(anly.num2alpha(num+10))
                            ws[ab+str(env.step_count+1)]=var[num]
                            if env.step_count>=1:
                                if ws[ab+str(env.step_count)].value==var[num] :
                                 ws[ab+str(env.step_count+1)].fill=fill                      
                    
                        print((str(env.step_count)+' steps completed').rjust(5, ' ')
                            ,' violation',str(np.round(env.violation,3)).rjust(6, ' ')
                            ,' volume',str(env.V).rjust(6, ' ')
                            ,'R',str(np.round(total_reward_t,2)).rjust(5, ' ') 
                            ,'CG:'+str(tnum_el[0])+str(tnum_el[1])+str(tnum_el[2])+'A:'+str(pact),'   ',np.max(Qs),' ',np.max(pacts),sign)

                            
                
                stop=time.time()
                print('Elapsed time',np.round(stop-start,2))
                ws['G'+str(1)]=np.round(stop-start,2)
                wb.save('./output_images/RL_'+str(env.name)+'.xlsx') 
                
                print(env.memberz[env.margin:-env.margin,env.margin:-env.margin,env.margin:-env.margin])
                
                print('Stress ratio')
                for ii in range(env.zSpan):
                    print(np.round(env.stress_ratio[ii,:,:],5))
                print('COFX')
                for ii in range(env.zSpan):
                    print(np.round(env.cofX[ii,:,:],2))
                print('COFY')
                for ii in range(env.zSpan):
                    print(np.round(env.cofY[ii,:,:],2))                    
                print('層間変形角')  
                print(np.round(env.deflect,2)) 
                cof=np.maximum(env.cofX,env.cofY)
                print('COF')
                for ii in range(env.zSpan):
                    print(np.round(cof[ii,:,:],2))                
                print(env.memberz) 


