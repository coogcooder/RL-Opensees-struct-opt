import gym
import gym.spaces
import numpy as np
import torch
from torch.autograd import Variable
import sys
import os
import pickle
sys.path.append("~/module")
###Original imports
from module import anly
from module import parameters as para
# import matlab.engine
import random
import openpyxl as excel
# import make_map
import copy

# filename1='./csv/column_list_build.csv' 
# filename2='./csv/beam_list_build.csv'
# col_list,list_c=anly.dctnry(filename1,0)
# beam_list,list_b=anly.dctnry(filename2,1)
# modelname='./csv/param_3d'

filename1='column_list_build.csv'
filename2='beam_list_build.csv'
col_list,list_c=anly.dctnry(filename1,0)
beam_list,list_b=anly.dctnry(filename2,1)
modelname='param_3d'


###Execute action
def dim_action(member,action,bc): ##Column/Beam (memberz)
  
  member=member[2:]

  if bc==1 :
   idz=anly.code2idz(member,0)
   D=int(idz[0])
   td=int(idz[1])
   member=anly.idz2code([D,td],0)
   if (action ==0 or action ==1)  :
      if  D==len(anly.list_D):
        pass
      else:
        member=anly.idz2code([D+1,td],0)
   elif (action ==2 ) :
      if td==len(anly.list_t):
        pass
      else:
        member=anly.idz2code([D,td+1],0)
   elif (action ==3 or action ==4) :
      if D==1:
        pass
      else:
        member=anly.idz2code([D-1,td],0)   
   elif (action ==5) :
      if td==1:
        pass
      else:
        member=anly.idz2code([D,td-1],0)

  
  else:
    idz=anly.code2idz(member,1)
    H=int(idz[0])
    B=int(idz[1])
    tw=int(idz[2])
    tf=int(idz[3])

    member=anly.idz2code([H,B,tw,tf],1)
    if   action ==0  :
      if H==len(anly.list_H):
        pass
      else:
        member=anly.idz2code([H+1,B,tw,tf],1)
    elif   action ==1 :
      if B==len(anly.list_B):
        pass
      else:
        member=anly.idz2code([H,B+1,tw,tf],1)
    elif   action ==2 :
      if tf==len(anly.list_tf):
        pass
      else:
        tw=np.argmin(abs(np.array(anly.list_tw)-anly.list_tf[tf+1-1]/2))+1
        member=anly.idz2code([H,B,tw,tf+1],1)  

    elif   action ==3  :
      if H==1 :
        pass
      else:
        member=anly.idz2code([H-1,B,tw,tf],1)    
    elif   action ==4   :
      if B==1 :
        pass
      else:
        member=anly.idz2code([H,B-1,tw,tf],1)
    elif   action ==5  :
      if tf==1 :
        pass
      else:
        tw=np.argmin(abs(np.array(anly.list_tw)-anly.list_tf[tf-1-1]/2))+1
        member=anly.idz2code([H,B,tw,tf-1],1)       
  
  
  return      member




class Env(gym.core.Env):
    '''
    You can change the number of sensors by modifying the observation_space
    You can change the number of possible movement directions by modifying the action_space
    You can change the target image for tracing by modifying the target_img_path
    '''

    def __init__(self):

        #Loop setting
        self.step_count = 0
        self.max_step= 50
        self.done=False   
          
        self.margin=2
        self.p=1.0  
        
        if os.path.isfile("./output_images/file.npy"):
          self.FRAME_list=np.load("./output_images/file.npy", allow_pickle=True)[()]
        else:  
          self.FRAME_list={'TEST':0}      


    def step(self, update_el,pact):   
        ##Variation range: Volume and Constraints
        self.num_el=self.sort[update_el]    
        self.result=[copy.copy(self.memberz),copy.copy(self.stress_ratio)
                     ,copy.copy(self.cofX),copy.copy(self.cofY),copy.copy(self.deflect),copy.copy(self.violation),copy.copy(self.V)]
        
        nxtcode=dim_action(self.memberz[self.num_el[0],self.num_el[1],self.num_el[2]]
                                    ,pact,self.arraymat[self.num_el[0],self.num_el[1],self.num_el[2]])
        self.memberz[self.num_el[0],self.num_el[1],self.num_el[2]] = nxtcode
        if pact<3:
            maxmin=0#Larger value
        else:
            maxmin=1#Smaller value    
          
        self.memberz=anly.mirror(self.memberz,maxmin,self.symFlag)
        self.flag=list(copy.copy(self.memberz).reshape((-1))) in self.frame_list
        # Structural analysis
        if not self.flag:
         stress_ratio,cofX,cofY,deflect=anly.r_cof(self.memberz,col_list,beam_list,self.margin,self.xSpan,self.ySpan,self.zSpan,self.arraymat,self.symFlag)    
         self.stress_ratio,self.cofX,self.cofY,self.deflect=copy.copy(stress_ratio),copy.copy(cofX),copy.copy(cofY),copy.copy(deflect)
         self.violation=anly.violate(self.stress_ratio,self.cofX,cofY,self.deflect,self.margin,self.arraymat)

        else:pass 
 
         
        self.V=anly.V(self.memberz,self.lengthmat,self.margin)
        dV=self.result[6]-self.V
        dviolation = self.result[5]-self.violation 
        
        
        reward = anly.reward_gvn(dV,self.result[5],dviolation)  
        if self.violation==0:reward=10*reward
        self.F=anly.update_flag(dV,self.result[5],dviolation)  
        if 1:#self.test_train=='train' or self.test_train=='test':
          if self.F==False or self.flag==True:          
            self.flag_counter=self.flag_counter+1
        else:   ##Verification
          if self.F==False or self.flag==True:
            self.memberz,self.stress_ratio,self.cofX,self.cofY,self.deflect,self.violation,self.V=(
              self.result[0],self.result[1],self.result[2],self.result[3],self.result[4],self.result[5],self.result[6])
            self.flag_counter=self.flag_counter+1
          else:
            self.flag_counter=0

        if   ( self.step_count==self.max_step ) and self.test_train=='train': 
            del self.frame_list
            self.done=True       
            
  
        else:
            self.step_count += 1  
            self.frame_list.append(list(copy.copy(self.memberz).reshape((-1))))
        return  reward, self.done,self.violation,dviolation,pact

    # Initialize environment and return state
    def reset(self,test_train):
        self.test_train=test_train
        if test_train=='train':
         #Determine the shape of the model
         self.xSpan=random.randint(2,3)
         self.ySpan=copy.copy(self.xSpan)
         self.zSpan=random.randint(3,4)             
         self.zlength=np.ones((self.zSpan))*4000
         self.zlength[0]=4000
         self.xlength=0
         self.load=0
         self.xlength=random.randint(12,20)*500
         self.ylength=copy.copy(self.xlength)
         self.load=random.randint(70,75)*1+7           
        #  while self.xlength/1000*self.load<30:
        #     self.xlength=random.randint(12,20)*500
        #     self.ylength=copy.copy(self.xlength)
        #     self.load=random.randint(0,15)*0.2+7  
        elif test_train=='test': 
        #  self.FRAME_list.pop('2x10.0_2y10.0_z4')
         #Determine the shape of the model
         self.xSpan=5
         self.ySpan=10
         self.zSpan=4
         self.xlength=4*2000  
         self.ylength=copy.copy(self.xlength)
         self.zlength=np.ones((self.zSpan))*4000      
         self.zlength[0]=4000  
         self.load=10.0
        else:
         self.xSpan=test_train[0]
         self.ySpan=test_train[1]
         self.zSpan=test_train[2]
         self.xlength=test_train[3]
         self.ylength=test_train[4]
         self.zlength=np.ones((self.zSpan))*4000
         self.load=float(test_train[5])
        self.lx=2*self.xSpan+1
        self.ly=2*self.ySpan+1
        #Write
        wb = excel.Workbook()
        ws = wb.active
        ws['A1']=self.xSpan
        ws['B1']=self.ySpan
        ws['C1']=self.zSpan
        ws['D1']=self.xlength
        ws['E1']=self.ylength
        ws['F1']=4000#zheight
        ws['G1']=self.load/1000#Load
        ws['H1']=2      
        for i in range(self.zSpan):
          code=anly.num2alpha(i+1) 
          ws[str(code)+"2"]=self.zlength[i] 
        self.model=[self.zSpan,self.ySpan,self.xSpan,4,self.ylength/1000,self.xlength/1000]           
        wb.save(modelname+'.xlsx')
        self.name=str(self.xSpan)+'x'+str(self.xlength/1000)+'_'+str(self.ySpan)+'y'+str(self.ylength/1000)+'_z'+str(self.zSpan)+'_W'+str(self.load)#Model identifier
    
        arraymat,self.arraymat=self.framemat()
        self.symFlag=self.symmetryFlag()
        if self.name in self.FRAME_list:
          self.memberz= copy.copy(self.FRAME_list[self.name])
        else:  
         if self.symFlag==1:
          vecFrame='x'
          self.memberz=anly.FRAME_gen(self.model,arraymat,self.margin,vecFrame,self.symFlag)
         else:
          vecFrameX='x'
          vecFrameY='y'
          self.memberz=anly.FRAME_gen(self.model,arraymat,self.margin,vecFrameX,self.symFlag)  
          memberY=     anly.FRAME_gen(self.model,arraymat,self.margin,vecFrameY,self.symFlag)
          for i in range(self.zSpan+2*self.margin):
            for j in range(self.ly+2*self.margin):
              for k in range(self.lx+2*self.margin):
                # memberY has dimensions (z, lx+2*margin, ly+2*margin)
                if self.memberz[i,j,k] <= memberY[i,k,j]:
                  self.memberz[i,j,k] = memberY[i,k,j]
                else:
                  pass
         self.FRAME_list[self.name] = copy.copy(self.memberz)  
         np.save('./output_images/file.npy', self.FRAME_list)
        self.lengthmat= self.flengthmat(arraymat)   
        self.step_count=1
        self.flag_counter=0
        self.flag=False
        self.done=False
        self.p=1
        self.init_V=anly.V(self.memberz,self.lengthmat,self.margin)#Initial volume calculation
        
        
        # Structural analysis
        stress_ratio,cofX,cofY,deflect=anly.r_cof(self.memberz,col_list,beam_list,self.margin,self.xSpan,self.ySpan,self.zSpan,self.arraymat,self.symFlag)
        self.violation=anly.violate(stress_ratio,cofX,cofY,deflect,self.margin,self.arraymat)      
        self.violation0=np.max([0.000001,self.violation])
        self.stress_ratio,self.cofX,self.cofY,self.deflect=copy.copy(stress_ratio),copy.copy(cofX),copy.copy(cofY),copy.copy(deflect)
        self.V=copy.copy(self.init_V)
        self.Vsave=copy.copy(self.init_V)
        self.sort=self.sort_el()
        self.frame_list=[list(copy.copy(self.memberz).reshape((-1)))]
        #cmd visual
        self.visual_frame=np.full(self.zSpan*self.ly*self.lx,'×',dtype=object).reshape((self.zSpan,self.ly,self.lx))
        for i in range(self.zSpan):
          for j in range(self.ly):
            for k in range(self.lx):
              if (j+1)%2==0 and (k+1)%2==0:
               self.visual_frame[i,j,k]=' '
        self.target_num=0       
        self. Q_check=copy.copy(self.arraymat)*0      
        self.resultb=[copy.copy(self.memberz),copy.copy(self.stress_ratio)
                      ,copy.copy(self.cofX),copy.copy(self.cofY),copy.copy(self.deflect),copy.copy(self.violation),copy.copy(self.V)]
  






                    


    def update(self,member):
      memberz=copy.copy(member)
      stress_ratio,cofX,cofY,deflect=anly.r_cof(memberz,col_list,beam_list,self.margin,self.xSpan,self.ySpan,self.zSpan,self.arraymat,self.symFlag)
      
      #Evaluation value after update
      V=anly.V(memberz,self.lengthmat,self.margin)
      violation=anly.violate(stress_ratio,cofX,cofY,deflect,self.margin,self.arraymat)
      PF=V/self.init_V+violation/self.violation0*self.p
      return PF

    



    def LSexplore(self,num_el,member0):#Member section update action - Select optimal member by LS
      member=copy.copy(member0)
      aply_act=0
      sw=0
      code=member0[num_el[0],num_el[1],num_el[2]]
      acts=self.dim_assets(num_el,code)
      
      while sw<3:
          Ms=[]
          pacts=[]
          F=[]
          PF0=self.update(copy.copy(member0))
          for action in acts:
            P_= self.update(member0)
            member_=copy.copy(member0)            
            nxtcode=dim_action(member0[num_el[0],num_el[1],num_el[2]]
                                        ,action,self.arraymat[num_el[0],num_el[1],num_el[2]])
            member_[num_el[0],num_el[1],num_el[2]]=nxtcode
            if action<3:
                maxmin=0#Larger value
            else:
                maxmin=1#Smaller value    
            
            member_=anly.mirror(member_,maxmin,self.symFlag)  
            PF=self.update(member_)
            if PF<PF0:
              member=copy.copy(member_)
              aply_act=action
              Ms.append(copy.copy(member_))
              pacts.append(action)
              F.append(PF)
              sw=5
          if sw<3:
             sw=sw+1
             if  self.violation==0:sw=5
             else: self.p=self.p*2    
      if bool(pacts):  
       if self.test_train=='test': id=np.argmin(F)      
       else:  id=random.randint(0,len(pacts)-1)           
       member=copy.copy(Ms[id])
       aply_act=pacts[id] 
      self.p=1.0
      return member,aply_act
   
    def sort_el(self):#Member reordering action
      sort=[]
      size1=np.shape(self.arraymat)
      for i in reversed(range(size1[0])) :
        for j in range(int((size1[1]+1)/2)) :
          for k in range(int((size1[2]+1)/2)) :
            if  self.arraymat[i,j,k]==2:
             if self.symFlag==1:
               if  k>=j :        
                 sort.append([i,j,k])    
             else:  
               sort.append([i,j,k])  
            if self.arraymat[i,j,k]==1:
             if self.symFlag==1:
               if  k>=j :        
                 sort.append([i,j,k])    
             else:  
               sort.append([i,j,k])  
                     
      return sort
    
    
    def value_scan_assets(self,NN,device):
      state=[]
      for i in range(len(self.sort)):
        s=anly.state_update(copy.copy(self.stress_ratio),copy.copy(self.cofX),copy.copy(self.cofY)
                                ,copy.copy(self.deflect),self.sort[i],self.margin,self.arraymat,self.memberz)
        state.append(torch.from_numpy(np.array(s, dtype="float32").reshape((1, para.obs_num1)) )  )

      state=torch.stack(state)
      pobs_=state.to(device)
      pact_=NN(pobs_).to('cpu').detach().numpy().copy()
      return np.array(pact_).reshape((-1))  

    def value_scan_assets_DIM(self,NN,device):
      Qs1=np.ones((len(self.sort),6))*(-100)
      for i in range(len(self.sort)): 
        num_el=self.sort[i]
        list_dim=self.dim_assets(num_el,self.memberz[num_el[0],num_el[1],num_el[2]])  
        state=anly.state_updateDIM(copy.copy(self.stress_ratio),copy.copy(self.cofX),copy.copy(self.cofY)
                                ,copy.copy(self.deflect),self.sort[i],self.margin,self.arraymat,self.memberz)           
        pobs_ = np.array(state, dtype="float32").reshape((1, 9))
        pobs_ = Variable(torch.from_numpy(pobs_))
        pobs_=pobs_.to(device)
        pacts_ = NN(pobs_).to('cpu').detach().numpy().copy()[0]
        for j in list_dim:
          Qs1[i,j]=pacts_[j]
      return np.array(Qs1)  


    def dim_assets(self,num_el,code):
      dim=anly.code2idz(str(code)[2:],self.arraymat[num_el[0],num_el[1],num_el[2]]-1)
      if self.arraymat[num_el[0],num_el[1],num_el[2]]==1:
        if self.violation>0:
          list=[0,2,3,5]
          if dim[0]==1:list.remove(3)
          elif dim[0]==13:list.remove(0)
          if dim[1]==1:list.remove(5)
          elif dim[1]==9:list.remove(2)          
        else:               
          list=[3,5]
          if dim[0]==1:list.remove(3)
          if dim[1]==1:list.remove(5)
      else:                                    
        if self.violation>0:
          list=[0,1,2,3,4,5]
          if dim[0]==1:list.remove(3)
          elif dim[0]==13:list.remove(0)
          if dim[1]==1:list.remove(4)
          elif dim[1]==4:list.remove(1)
          if dim[3]==1:list.remove(5)
          elif dim[3]==8:list.remove(2)        
        else:               
          list=[3,4,5]
          if dim[0]==1:list.remove(3)
          if dim[1]==1:list.remove(4)
          if dim[3]==1:list.remove(5)

      return list

    def dim_assets_(self,num_el,code):
      if len(str(code)[2:])>2:cg=2
      else:cg=1
      dim=anly.code2idz(str(code)[2:],cg)
      if cg==1:
        list=[0,2,3,5]
       
        if dim[0]==1:list.remove(3)
        elif dim[0]==13:list.remove(0)
        if dim[1]==1:list.remove(5)
        elif dim[1]==9:list.remove(2)
      else:                                    
        list=[0,1,2,3,4,5]
        
        if dim[0]==1:list.remove(3)
        elif dim[0]==13:list.remove(0)
        if dim[1]==1:list.remove(4)
        elif dim[1]==4:list.remove(1)
        if dim[3]==1:list.remove(5)
        elif dim[3]==8:list.remove(2)

      return list   
    
    def framemat(self):
      arraymat=np.ones((self.zSpan,self.ly,self.lx))*2
      arraymat_margin=np.zeros((self.zSpan+2*self.margin,self.ly+2*self.margin,self.lx+2*self.margin))
      for i in range(self.zSpan):
        for j in range(self.ly):
          for k in range(self.lx):
            if j%2==0 and k%2==0:
              arraymat[i,j,k]=1
            elif   j%2==1 and k%2==1:
              arraymat[i,j,k]=0
      arraymat_margin[self.margin:-self.margin,self.margin:-self.margin,self.margin:-self.margin]=arraymat
      return arraymat  ,arraymat_margin       


    def flengthmat(self,arraymat):#Length matrix calculation
      lengthmat=np.zeros((self.zSpan,self.ly,self.lx))
      for i in range(self.zSpan):
        for j in range(self.ly):
          for k in range(self.lx):
            if arraymat[i,j,k]==1:
              lengthmat[i,j,k]=self.zlength[i]
            if arraymat[i,j,k]==2:
              if k%2==1:#ｘ方向
               lengthmat[i,j,k]=self.xlength
              else: #ｙ方向
                lengthmat[i,j,k]=self.ylength

      return lengthmat 

    def symmetryFlag(self):#Symmetry check
      if self.xSpan==self.ySpan and self.xlength==self.ylength:
        flag=1
      else:
        flag=0  
      return flag


