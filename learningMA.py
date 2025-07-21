import copy
import time
import os
import numpy as np
import matplotlib.pyplot as plt
import torch
import torch.nn as nn
import torch.nn.functional as F
import torch.optim as optim
from torch.autograd import Variable
import openpyxl as excel
import traceback
import sys
sys.path.append("~/module")
###Original imports
from module import envmnt_lrn
from module import anly
from module import visualize
from module import NN
from module import parameters as para

####GPU
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
print(device)
###Environment expansion
env = envmnt_lrn.Env()
EPSILON1,EPSILON2=para.EPSILON1,para.EPSILON2

try:
 for rep in range(1):

    # Constants


    # Model
    Qel = NN.NN1(para.obs_num1,para.acts_num1).to(device)   
    Qdim= NN.NN2(para.obs_num2,para.acts_num2).to(device)                
    Qel_ast = copy.deepcopy(Qel)  
    Qdim_ast = copy.deepcopy(Qdim)  

    optimizer_dim = optim.Adam(Qdim.parameters(), lr=0.0001, betas=(0.9, 0.999), eps=1e-04, weight_decay=0, amsgrad=False)
    optimizer_el = optim.Adam(Qel.parameters(), lr=0.0001, betas=(0.9, 0.999), eps=1e-04, weight_decay=0, amsgrad=False)

    data1,data2,data3,data4,data5 = [],[] ,[] ,[] ,[]    
    memory_el = []                                 
    total_rewards = [] 
    total_rewards_t = []
    reward_history=[]   
    total_volume_t=[] 
    plt1,plt2,plt3=[],[],[]

    step=0




    ##Learning process
    print("●Learning Start")
    start = time.time()
    wb=excel.Workbook()
    ws=wb['Sheet']
    fname='learning_data'
    id_fname=0
    while os.path.exists('./output_images/'+str(fname)+str(id_fname)):
        id_fname=id_fname+1
    fname='./output_images/'+str(fname)+str(id_fname)
    os.mkdir(fname)

    # for epoch in range(para.EPOCH_NUM):
    #     env.reset('train')                         # Environment initialization
    #     done = False                               # Game end flag
    for epoch in range(para.EPOCH_NUM):
        env.reset('train')                         # Environment initialization
        # if epoch == 0:
        #     ncol = (env.xSpan + 1) * (env.ySpan + 1) * env.zSpan
        #     nbx = env.xSpan * (env.ySpan + 1) * env.zSpan
        #     nby = env.ySpan * (env.xSpan + 1) * env.zSpan
            # print(f"Columns: {ncol}, Beams-X: {nbx}, Beams-Y: {nby}")
            # print('Stress ratio')
            # print(env.stress_ratio)
            # print('COFX')
            # print(env.cofX)
            # print('COFY')
            # print(env.cofY)
            # print('Deflection')
            # print(env.deflect)
            # sys.exit(0)
        done = False                               # Game end flag
        total_reward = 0                           # Cumulative reward
        list_dim=list(np.arange(para.acts_num2))
        opt,rand=0,0
        env.switch=(epoch)%2 
        while not done :
            if env.switch==1:
                array=env.value_scan_assets_DIM(Qdim,device)

                if env.flag_counter>0:
                    for i in range(env.flag_counter):
                        index = np.unravel_index(np.argmax(array), array.shape) 
                        array[index[0],index[1]] =-100
                index = np.unravel_index(np.argmax(array), array.shape) 
                pact=index[0]
                pact2=index[1]
                num_el=env.sort[pact]
                list_dim=env.dim_assets(num_el,env.memberz[num_el[0],num_el[1],num_el[2]])
                thred=np.random.rand()
                if thred > EPSILON1:
                    opt=opt+1
                else: #Column:1 Beam:2
                    array_act=np.arange(array.size).reshape((array.shape))
                    if np.max(array)>-100:
                        a = np.random.choice(array_act[array>-100]) 
                        ids=np.where(array_act==a)
                        pact,pact2=ids[0][0],ids[1][0]
                    else:
                        a = np.random.choice(array_act[array>=-100]) 
                        ids=np.where(array_act==a)
                        pact,pact2=ids[0][0],ids[1][0]
                    rand=rand+1
            else:
                list_pact=list(np.arange(len(env.sort)))
                Qs=list(env.value_scan_assets(Qel,device)) 
                if env.flag_counter>0:
                    for i in range(env.flag_counter):
                        index = np.argmax(Qs)
                        Qs[index] =-100                  
                    
                index = list_pact[np.argmax(Qs)]
                thred=np.random.rand()
           
                if thred > EPSILON2:
                    # Predict the optimal action
                    pact = index
                    opt=opt+1
                else:
                    list_pact=np.arange(len(env.sort))
                    if np.max(np.array(Qs))>-100:
                        pact = np.random.choice(list_pact[np.array(Qs)>-100]) 
                    else:
                        pact = np.random.choice(list_pact)
                    env.flag_counter=0
                    rand=rand+1   
                # Action
                num_el=env.sort[pact]
                list_dim=env.dim_assets(num_el,env.memberz[num_el[0],num_el[1],num_el[2]])                
                if len(list_dim)>0:
                    # Predict the optimal action
                    pobs = anly.state_updateDIM(copy.copy(env.stress_ratio),copy.copy(env.cofX),copy.copy(env.cofY)
                                            ,copy.copy(env.deflect),num_el,env.margin,env.arraymat,env.memberz)                    
                    pobs_ = np.array(pobs, dtype="float32").reshape((1, para.obs_num2))
                    pobs_ = torch.from_numpy(pobs_).to(device)
                    pacts = Qdim(pobs_).to('cpu').detach().numpy().copy()[0]
                    pacts=pacts[list_dim]
                    id=np.argmax(pacts)
                    pact2=list_dim[id] 
                else:
                    _,pact2 = env.LSexplore(num_el,env.memberz)


            pS_dim = anly.state_updateDIM(copy.copy(env.stress_ratio),copy.copy(env.cofX),copy.copy(env.cofY)
                                        ,copy.copy(env.deflect),num_el,env.margin,env.arraymat,env.memberz)
            pS_el = anly.state_update(copy.copy(env.stress_ratio),copy.copy(env.cofX),copy.copy(env.cofY)
                                        ,copy.copy(env.deflect),num_el,env.margin,env.arraymat,env.memberz)                
            pS_dim = torch.from_numpy(np.array(pS_dim, dtype="float32").reshape((1, para.obs_num2))).to(device)
            pS_el = torch.from_numpy(np.array(pS_el, dtype="float32").reshape((1, para.obs_num1))).to(device)
                # Action
            reward, done,violation,_1,pact2 = env.step(pact,pact2) 
            
            # Memory accumulation
            Qs=list(env.value_scan_assets(Qel,device))  
            list_pact=list(np.arange(len(env.sort)))      
            index = list_pact[np.argmax(Qs)]  
            num_el=env.sort[index]              
            S_dim = anly.state_updateDIM(copy.copy(env.stress_ratio),copy.copy(env.cofX),copy.copy(env.cofY)
                                        ,copy.copy(env.deflect),num_el,env.margin,env.arraymat,env.memberz)
            S_el = anly.state_update(copy.copy(env.stress_ratio),copy.copy(env.cofX),copy.copy(env.cofY)
                                        ,copy.copy(env.deflect),num_el,env.margin,env.arraymat,env.memberz) 
            S_dim = torch.from_numpy(np.array(S_dim, dtype="float32").reshape((1, para.obs_num2))).to(device)
            S_el = torch.from_numpy(np.array(S_el, dtype="float32").reshape((1, para.obs_num1))).to(device)                  
            


            if env.F==False or env.flag==True:
                done_=False
                env.memberz,env.stress_ratio,env.cofX,env.cofY,env.deflect,env.violation,env.V=(
                    copy.copy(env.result[0]),copy.copy(env.result[1]),copy.copy(env.result[2]),copy.copy(env.result[3])
                            ,copy.copy(env.result[4]),copy.copy(env.result[5]),copy.copy(env.result[6]))
            else:
                env.flag_counter=0
                done_=False              
            total_reward += reward        
            step=step+1

            data1.append(pS_dim.data.to('cpu')),data2.append(S_dim.data.to('cpu'))
            data3.append(pS_el.data.to('cpu')),data4.append(S_el.data.to('cpu'))
            data5.append([pact2, reward, done_,env.memberz[env.num_el[0],env.num_el[1],env.num_el[2]],env.num_el])
            if len(data3) > para.MEMORY_SIZE:                  # メモリサイズを超えていれば消していく
               data1.pop(0),data2.pop(0),data3.pop(0),data4.pop(0),data5.pop(0)
            
            if len(data3) >= para.BATCH_SIZE: # メモリサイズ分溜まっていれば学習
            # 経験リプレイ
              if (epoch+1) % para.TRAIN_FREQ == 0:
        
                dataset=NN.MyDataset(data1,data2,data3,data4,data5,list(range(len(data1)))) 
                trainloader=torch.utils.data.DataLoader(dataset, batch_size=para.BATCH_SIZE, shuffle=True)
                for i, (pSd,Sd,pSe,Se,assets,sample) in enumerate(trainloader,0): 
                        if i>0:break
                        else:
                            ##assets
                            pacts=assets[0].tolist()
                            rs=assets[1].tolist()
                            dones=assets[2].tolist()                        
                            codes=assets[3]
                            num_els=assets[4]
                            z,y,x=num_els[0].tolist(),num_els[1].tolist(),num_els[2].tolist()   

                        #y
                            pSd=pSd.clone().detach().to(device)
                            pSe=pSe.clone().detach().to(device)
                            Sd=Sd.clone().detach().to(device)
                            Se=Se.clone().detach().to(device)
                            qel = Qel(pSe).squeeze()
                            qel_ = Qel_ast(Se).squeeze()
                            target = copy.deepcopy(qel.data.to('cpu').detach().numpy().copy())
                            maxqs = copy.deepcopy(qel_.data.to('cpu').detach().numpy().copy())
                            for j in range(para.BATCH_SIZE):
                                target[j] = rs[j]+para.GAMMA*maxqs[j]*(1-dones[j]) # 教師信号

                            qel=qel.to(device)
                            labels=torch.from_numpy(target).to(device)
                            optimizer_el.zero_grad()
                            loss = nn.MSELoss()(qel, labels)
                            loss.backward()
                            optimizer_el.step()
                            
                            ##断面選択
                            qdim = Qdim_ast(pSd).squeeze()
                            qdim_ = Qdim_ast(Sd).squeeze()
                            target = copy.deepcopy(qdim.data.to('cpu').detach().numpy().copy())
                            maxqs = copy.deepcopy(qdim_.data.to('cpu').detach().numpy().copy())                       
                                    
                            for j in range(para.BATCH_SIZE):
                                acts=maxqs[j, :].squeeze()
                                num_el=[z[j],y[j],x[j]]
                                list_dim=env.dim_assets_(num_el,codes[j])
                                acts=acts[list_dim]
                                id=list_dim[np.argmax(acts)]                                
                                target[j, pacts[j]] = rs[j]+para.GAMMA*maxqs[j,id]*(1-dones[j]) # 教師信号  
                            #  Perform a gradient descent step
                            qdim=qdim.to(device)
                            labels=torch.from_numpy(target).to(device)
                            optimizer_dim.zero_grad()
                            loss = nn.MSELoss()(qdim, labels)
                            loss.backward()
                            optimizer_dim.step()


                     

                # Q-function update
                if step % para.UPDATE_TARGET_Q_FREQ == 0:
                    Qdim_ast = copy.deepcopy(Qdim)
                    Qel_ast = copy.deepcopy(Qel)


        
        if EPSILON1 > para.EPSILON_MIN  and env.switch==1:
            EPSILON1 = EPSILON1 - para.EPSILON_DECREASE  
        if EPSILON2 > para.EPSILON_MIN and env.switch==0:
            EPSILON2 = EPSILON2 - para.EPSILON_DECREASE         
        


        total_rewards.append(total_reward) # 累積報酬を記録
        ws['A'+str(epoch+1)]=env.name
        ws['B'+str(epoch+1)]=epoch+1
        ws['C'+str(epoch+1)]=total_reward
        ws['D'+str(epoch+1)]=env.V
        ws['E'+str(epoch+1)]=env.violation   
        # print('終了',total_reward)
        print('□EP'+str(env.switch),str(epoch+1).rjust(5, ' '),(env.name).ljust(20, ' '),'O/R',str(opt).rjust(2, ' '),'/',str(rand).ljust(2, ' ')
            ,' violation',str(np.round(env.violation,4)).rjust(6, ' '),'←',str(np.round(env.violation0,4)).ljust(6, ' ')
            ,' volume',str(env.V).rjust(7, ' '),'←',str(env.init_V).ljust(7, ' ')
            ,'R',str(np.round(total_reward,4)).rjust(5, ' '))
        ####エージェントのテスト
        #部材履歴保存
        if (epoch+1) % (10) == 0:
            print('                                ','□EP'+str(env.switch),str(epoch+1).rjust(5, ' '))
                            
            done_t = False                           
            total_reward_t = 0     #      
            env.reset('test') 
            print((str(env.step_count)+' steps completed').rjust(5, ' ')
                ,' violation',str(np.round(env.violation,4)).rjust(6, ' '),' volume',str(env.V).rjust(6, ' ')) 
            while not done_t:
                Qs=env.value_scan_assets(Qel,device)
                if env.flag_counter>0:
                    for i in range(env.flag_counter):
                        index = np.argmax(Qs)
                        Qs[index] =-100
                index = np.argmax(Qs)        
                if  env.step_count==env.max_step or np.max(Qs)<-90:break
             
                index = np.argmax(Qs)
                tnum_el=env.sort[index]
                update_pos = index
                pobs = anly.state_updateDIM(copy.copy(env.stress_ratio),copy.copy(env.cofX),copy.copy(env.cofY)
                                            ,copy.copy(env.deflect),tnum_el,env.margin,env.arraymat,env.memberz)
                # list_dim=env.dim_assets(env.sort[index],env.memberz[tnum_el[0],tnum_el[1],tnum_el[2]])
                #     # 最適な行動を予測
                # pobs_ = np.array(pobs, dtype="float32").reshape((1, para.obs_num2))
                # pobs_ = torch.from_numpy(pobs_).to(device)
                # pacts = Qdim(pobs_).to('cpu').detach().numpy().copy()[0]
                # pacts=pacts[list_dim]
                # id=np.argmax(pacts)
                # pact=list_dim[id] 
                list_dim=env.dim_assets(env.sort[index],env.memberz[tnum_el[0],tnum_el[1],tnum_el[2]])
                if len(list_dim)==0:
                    env.flag_counter += 1
                    continue
                # 最適な行動を予測
                pobs_ = np.array(pobs, dtype="float32").reshape((1, para.obs_num2))
                pobs_ = torch.from_numpy(pobs_).to(device)
                pacts = Qdim(pobs_).to('cpu').detach().numpy().copy()[0]
                pacts=pacts[list_dim]
                id=np.argmax(pacts)
                pact=list_dim[id]
                if 0:pass#np.max(pacts)<0:env.flag_counter +=1
                else:
                    # 行動 
                    reward, done_t,member,dviolation,pact= env.step(update_pos,pact) 
                    total_reward_t += max([0,reward])    
                 
                    
                    sign='●'
                    if env.F==False or env.flag==True:
                        sign='　'    
                        env.memberz,env.stress_ratio,env.cofX,env.cofY,env.deflect,env.violation,env.V=(
                            copy.copy(env.result[0]),copy.copy(env.result[1]),copy.copy(env.result[2]),copy.copy(env.result[3])
                                    ,copy.copy(env.result[4]),copy.copy(env.result[5]),copy.copy(env.result[6]))               
                    else:env.flag_counter=0

                    print((str(env.step_count)+' steps completed').rjust(5, ' ')
                    ,' violation',str(np.round(env.violation,4)).rjust(6, ' '),' volume',str(env.V).rjust(6, ' ')
                    ,'R',str(np.round(total_reward_t,4)).rjust(6, ' ') 
                    ,'CG:'+str(tnum_el[0])+str(tnum_el[1])+str(tnum_el[2])+'A:'+str(pact)
                    ,sign,' ',np.max(Qs))
                
            total_rewards_t.append(total_reward_t)  
            reward_history.append(total_reward_t)
            print(env.memberz[env.margin:-env.margin,env.margin:-env.margin,env.margin:-env.margin])
            print(env.stress_ratio)
            print(env.deflect)
            print(env.cofX)
            print(env.cofY)
            plt1.append(env.V)
            if env.violation==0:plt2.append(len(plt1)-1),plt3.append(env.V)             
            
            ws['F'+str(epoch+1)]=env.name
            ws['G'+str(epoch+1)]=total_reward_t
            ws['H'+str(epoch+1)]=env.V
            ws['I'+str(epoch+1)]=env.violation  

            
            if env.violation==0 and np.max(total_rewards_t)<=total_reward_t:
                        PATH=fname+'/agentDIM'+str(epoch+1)+'.pth'
                        torch.save(Qdim.state_dict(), PATH)
                        PATH=fname+'/agentEL'+str(epoch+1)+'.pth'
                        torch.save(Qel.state_dict(), PATH)    
            if  env.violation==0 :
                        PATH=fname+'/_agentDIM'+str(epoch+1)+'.pth'
                        torch.save(Qdim.state_dict(), PATH)
                        PATH=fname+'/_agentEL'+str(epoch+1)+'.pth'
                        torch.save(Qel.state_dict(), PATH)    

        if (epoch+1) % 50 == 0:
            np.save(fname+'/file.npy', env.FRAME_list)
            visualize.rwrd_plot(total_rewards,'_learning',1)
            visualize.rwrd_plot(total_rewards_t,'_test',plt_intrvl=para.LOG_FREQ)
            fig, ax1 = plt.subplots(figsize=(10, 5))
            # 左側の縦軸
            ax1.set_xlabel('epoch')
            ax1.set_ylabel('volume')
            ax1.plot(np.arange(len(plt1)), plt1, color='red', linestyle='-')  
            if len(plt2)>0:
                ax1.plot(plt2 , plt3, '*',color="b",markersize=10)
            ax1.tick_params(axis='y')
            plt.grid(True)
            plt.savefig(fname+f'_test.jpg')
            plt.close() 


            ws['J'+str(1)]='Calculation time'
            ws['J'+str(2)]=time.time()-start
            wb.save(fname+'/reward_history.xlsx')  
            wb = excel.load_workbook(fname+"/reward_history.xlsx")
            ws = wb['Sheet']


    PATH=fname+'/agentDIM'+str(epoch+1)+'.pth'
    torch.save(Qdim.state_dict(), PATH)
    PATH=fname+'/agentEL'+str(epoch+1)+'.pth'
    torch.save(Qel.state_dict(), PATH)      
    visualize.rwrd_plot(total_rewards_t,'_test',plt_intrvl=para.LOG_FREQ)  
    np.save(fname+'/file.npy', env.FRAME_list)    
    stop = time.time()
    elapsed_time = stop-start
    print('Elapsed time : '+str(elapsed_time)+' sec')
    ws['J'+str(1)]='Calculation time'
    ws['J'+str(2)]=elapsed_time
    wb.save(fname+'/reward_history.xlsx') 

except:
    traceback.print_exc()
    np.save(fname+'/file.npy', env.FRAME_list)
    visualize.rwrd_plot(total_rewards_t,'_test',plt_intrvl=para.LOG_FREQ)      
    stop = time.time()
    elapsed_time = stop-start
    print('Elapsed time : '+str(elapsed_time)+' sec')
    ws['J'+str(1)]='Calculation time'
    ws['J'+str(2)]=elapsed_time
    wb.save(fname+'/reward_history.xlsx') 

    PATH=fname+'/agentDIM'+str(epoch+1)+'terminated.pth'
    torch.save(Qdim.state_dict(), PATH)

    PATH=fname+'/agentEL'+str(epoch+1)+'terminated.pth'
    torch.save(Qel.state_dict(), PATH)           



